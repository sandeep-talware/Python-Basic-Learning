import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_worksheet(filename):
    wb = xl.load_workbook(filename)
    sh = wb['Sheet1']
    for row in range(2,sh.max_row+1):
        cell = sh.cell(row,3)
        correct_price = cell.value * 0.9
        correct_price_cell = sh.cell(row,4)
        correct_price_cell.value = correct_price
    values = Reference(sh,min_row=2,max_row=sh.max_row,min_col=4,max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sh.add_chart(chart,"E2")
    wb.save(filename)



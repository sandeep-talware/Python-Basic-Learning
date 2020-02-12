import openpyxl as xl

wb = xl.load_workbook("transactions.xlsx")
sh = wb['Sheet1']
cell = sh['a1']
print(cell.value)
cell = sh.cell(1,1)
print(cell.value)

import openpyxl as xl

wb = xl.load_workbook("transactions.xlsx")
sh = wb['Sheet1']
for row in range(2,sh.max_row+1):
    cell = sh.cell(row,3)
    correct_price = cell.value * 0.9
    correct_price_cell = sh.cell(row,4)
    correct_price_cell.value = correct_price
wb.save('tran2.xlsx')